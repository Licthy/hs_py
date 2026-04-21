"""
费用统计软件
读取 input.txt 和 num_name.txt，统计每日费用、总计，并生成 GUI 界面和 txt 报告
"""

import re
import json
import os
from datetime import datetime
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, asdict
import sys

# openpyxl 导入（Excel 导出）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# matplotlib 导入（可选，如果失败会给出提示）
try:
    import matplotlib.pyplot as plt
    # matplotlib 3.10+ 使用 backend_qtagg（兼容 PyQt5/6）
    try:
        from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    except ImportError:
        # 旧版本兼容
        try:
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        except ImportError:
            from matplotlib.backends.backend_qt6agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError as e:
    MATPLOTLIB_AVAILABLE = False
    print(f"警告：未找到 matplotlib，图表功能将不可用。错误：{e}")
    print("请安装：pip install matplotlib")

# 尝试导入 PyQt5 或 PyQt6
try:
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QTableWidget, QTableWidgetItem, QLabel, QPushButton, QMessageBox,
        QTabWidget, QTextEdit, QHeaderView, QFileDialog
    )
    from PyQt5.QtCore import Qt, QThread, QObject, pyqtSignal, pyqtSlot
    from PyQt5.QtGui import QFont
    PYQT_VERSION = 5
    # 别名
    AlignCenter = Qt.AlignCenter
    AlignLeft = Qt.AlignLeft
    AlignRight = Qt.AlignRight
except ImportError:
    try:
        from PyQt6.QtWidgets import (
            QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
            QTableWidget, QTableWidgetItem, QLabel, QPushButton, QMessageBox,
            QTabWidget, QTextEdit, QHeaderView, QFileDialog
        )
        from PyQt6.QtCore import Qt, QThread, QObject, pyqtSignal, pyqtSlot
        from PyQt6.QtGui import QFont
        PYQT_VERSION = 6
        # 别名
        AlignCenter = Qt.AlignmentFlag.AlignCenter
        AlignLeft = Qt.AlignmentFlag.AlignLeft
        AlignRight = Qt.AlignmentFlag.AlignRight
    except ImportError:
        print("错误：未找到 PyQt5 或 PyQt6。")
        print("请安装：pip install PyQt5  或  pip install PyQt6")
        sys.exit(1)


# ==================== 数据类定义 ====================

@dataclass
class DailyRecord:
    """单日费用记录"""
    date_str: str  # 如 "4-11"
    items: List[Tuple[int, float]]  # [(编号, 费用), ...]

@dataclass
class Statistics:
    """统计结果"""
    daily_data: Dict[str, Dict[int, float]]  # {日期: {编号: 费用}}
    num_totals: Dict[int, float]  # {编号: 总费用}
    daily_totals: Dict[str, float]  # {日期: 总费用}
    grand_total: float  # 总总计
    all_dates: List[str]  # 所有日期（按顺序）


# ==================== 数据解析器 ====================

class DataParser:
    """解析 input.txt 和 num_name.txt"""

    def __init__(self, input_path: str, mapping_path: str):
        self.input_path = Path(input_path)
        self.mapping_path = Path(mapping_path)
        self.num_to_name: Dict[int, str] = {}
        self.raw_records: List[DailyRecord] = []

    def load_num_mapping(self) -> Dict[int, str]:
        """加载编号-名称映射"""
        mapping = {}
        # 如果路径为空或文件不存在，返回空映射
        if not str(self.mapping_path).strip() or not self.mapping_path.exists() or not self.mapping_path.is_file():
            return mapping

        with open(self.mapping_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                # 格式：编号,名称 或 编号,名称,...
                parts = line.split(',')
                if len(parts) >= 2:
                    try:
                        num = int(parts[0].strip())
                        name = parts[1].strip()
                        mapping[num] = name
                    except ValueError:
                        continue
        return mapping

    def parse_input_file(self) -> Tuple[List[DailyRecord], List[str]]:
        """解析 input.txt
        Returns:
            (records, errors): 记录列表和错误信息列表
        """
        records = []
        errors = []
        # 检查路径是否有效：非空、存在且为文件
        if not str(self.input_path).strip():
            errors.append("未指定费用文件路径，请先选择文件。")
            return records, errors
        if not self.input_path.exists():
            errors.append(f"文件不存在：{self.input_path}")
            return records, errors
        if not self.input_path.is_file():
            errors.append(f"路径不是文件：{self.input_path}")
            return records, errors

        with open(self.input_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # 提取所有 {[日期],[费用列表]} 格式的数据块
        # 正则匹配：{[4,11],[...]} 或 {[4,12],[...]}
        pattern = r'\{\[(\d+),(\d+)\],\[(.*?)\]\}'
        matches = re.findall(pattern, content)

        line_num = 0
        for match in re.finditer(pattern, content):
            month, day, items_str = match.groups()
            line_num = content[:match.start()].count('\n') + 1
            date_str = f"{month}-{day}"

            # 验证月份和日期范围
            try:
                m = int(month)
                d = int(day)
                if not (1 <= m <= 12 and 1 <= d <= 31):
                    errors.append(f"第{line_num}行：日期格式错误 [{month},{day}]")
                    continue
            except ValueError:
                errors.append(f"第{line_num}行：日期必须是数字")
                continue

            items = []
            if items_str.strip():
                # 费用列表格式：编号,费用,编号,费用,...
                parts = [p.strip() for p in items_str.split(',') if p.strip()]
                if len(parts) % 2 != 0:
                    errors.append(f"第{line_num}行：费用项数量应为偶数（编号-费用对），实际 {len(parts)} 项")
                    # 继续处理，截断到偶数
                    parts = parts[:len(parts)//2*2]

                for i in range(0, len(parts), 2):
                    try:
                        num = int(parts[i])
                        cost = float(parts[i + 1])
                        if cost < 0:
                            errors.append(f"第{line_num}行：费用不能为负数（编号{num}）")
                            continue
                        items.append((num, cost))
                    except ValueError as e:
                        errors.append(f"第{line_num}行：数值格式错误 - {e}")
                        continue

            records.append(DailyRecord(date_str=date_str, items=items))

        return records, errors

    def parse(self) -> Tuple[List[DailyRecord], Dict[int, str], List[str]]:
        """执行解析
        Returns:
            (records, num_to_name, errors): 记录、映射、错误列表
        """
        self.num_to_name = self.load_num_mapping()
        self.raw_records, errors = self.parse_input_file()
        return self.raw_records, self.num_to_name, errors


# ==================== 统计引擎 ====================

class StatisticsEngine:
    """统计引擎"""

    @staticmethod
    def compute(records: List[DailyRecord]) -> Statistics:
        """计算统计数据"""
        # {日期: {编号: 费用}}
        daily_data = defaultdict(lambda: defaultdict(float))
        # {编号: 总费用}
        num_totals = defaultdict(float)
        # {日期: 总费用}
        daily_totals = defaultdict(float)
        all_dates_set = set()

        for record in records:
            date_str = record.date_str
            all_dates_set.add(date_str)
            for num, cost in record.items:
                daily_data[date_str][num] += cost
                num_totals[num] += cost
                daily_totals[date_str] += cost

        # 确保所有日期都在 daily_totals 中（即使空日期也显示为 0）
        for date in all_dates_set:
            if date not in daily_totals:
                daily_totals[date] = 0.0

        # 按日期排序
        all_dates = sorted(all_dates_set, key=lambda x: tuple(map(int, x.split('-'))))
        daily_data_dict = {date: dict(daily_data[date]) for date in all_dates}

        grand_total = sum(num_totals.values())

        return Statistics(
            daily_data=daily_data_dict,
            num_totals=dict(num_totals),
            daily_totals=dict(daily_totals),
            grand_total=grand_total,
            all_dates=all_dates
        )

    @staticmethod
    def generate_report(records: List[DailyRecord], stats: Statistics,
                        num_to_name: Dict[int, str]) -> str:
        """生成文本报告"""
        lines = []
        lines.append("=" * 60)
        lines.append("费用统计报告")
        lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 60)
        lines.append("")

        # ========== 1. 总总计（最上面） ==========
        lines.append("【总总计】")
        lines.append("-" * 60)
        lines.append(f"  全部费用合计：{stats.grand_total:.1f} 元")
        lines.append("")

        # ========== 2. 各编号费用总计 ==========
        lines.append("【各编号费用总计】")
        lines.append("-" * 60)
        for num in sorted(stats.num_totals.keys()):
            name = num_to_name.get(num, f"编号{num}")
            total = stats.num_totals[num]
            lines.append(f"  {num}（{name}）: {total:.1f} 元")
        lines.append("")

        # ========== 3. 每日费用总计 ==========
        lines.append("【每日费用总计】")
        lines.append("-" * 60)
        for date in stats.all_dates:
            lines.append(f"  {date}: {stats.daily_totals[date]:.1f} 元")
        lines.append("")

        # ========== 4. 各编号费用明细（最后） ==========
        lines.append("【各编号费用明细】")
        lines.append("-" * 60)
        # 按编号排序
        for num in sorted(stats.num_totals.keys()):
            name = num_to_name.get(num, f"编号{num}")
            lines.append(f"\n{num}（{name}）：")
            # 找出该编号有费用的所有日期
            has_data = False
            for date in stats.all_dates:
                cost = stats.daily_data[date].get(num, 0.0)
                if cost > 0:
                    lines.append(f"  {date}: {cost:.1f} 元")
                    has_data = True
            if not has_data:
                lines.append("  （无费用记录）")

        lines.append("")
        lines.append("=" * 60)
        lines.append("报告结束")
        lines.append("=" * 60)

        return "\n".join(lines)


# ==================== 异步数据加载 ====================

class DataLoaderWorker(QObject):
    """后台 worker：加载并解析数据，避免UI卡顿"""
    # 定义信号：finished(records, stats, num_to_name, errors, input_path, mapping_path)
    finished_signal = pyqtSignal(list, object, dict, list, str, str)
    error_signal = pyqtSignal(str)

    def __init__(self, input_path: str, mapping_path: str):
        super().__init__()
        self.input_path = input_path
        self.mapping_path = mapping_path

    @pyqtSlot()
    def run(self):
        """在后台线程中执行耗时操作"""
        try:
            # 解析文件
            parser = DataParser(self.input_path, self.mapping_path)
            records, num_to_name, errors = parser.parse()

            # 如果没有错误且有记录，计算统计
            stats = None
            if not errors and records:
                stats = StatisticsEngine.compute(records)

            # 发送完成信号
            self.finished_signal.emit(records, stats, num_to_name, errors, self.input_path, self.mapping_path)
        except Exception as e:
            # 发送错误信号
            self.error_signal.emit(str(e))


# ==================== PyQt5 GUI 界面 ====================

class StatisticsGUI(QMainWindow):
    """主窗口"""

    def __init__(self, records: List[DailyRecord], stats: Optional[Statistics],
                 num_to_name: Dict[int, str], errors: List[str] = None,
                 input_path: str = "input.txt", mapping_path: str = "num_name.txt"):
        super().__init__()
        self.records = records if records else []
        self.stats = stats
        self.num_to_name = num_to_name
        self.figure = None
        self.canvas = None
        self.errors = errors or []
        self.input_path = input_path
        self.mapping_path = mapping_path
        # 异步加载相关
        self.loading_thread = None
        self.loading_worker = None
        self.is_loading = False

        self.init_ui()

        # 根据是否有有效数据且无错误决定是否刷新界面内容
        if self.stats is not None and (self.stats.all_dates or self.stats.num_totals) and not self.errors:
            self.refresh_table()
            self.refresh_summary()
            self.refresh_raw_data()
            if MATPLOTLIB_AVAILABLE:
                self.refresh_chart()
        else:
            # 显示无数据状态（包括存在错误或没有有效数据）
            self.show_no_data_state()

        # 更新顶部摘要标签（始终调用，确保状态正确）
        self.refresh_summary_label()

        # 如果有错误，在启动时显示
        if self.errors:
            error_msg = "\n".join([f"• {err}" for err in self.errors[:10]])
            if len(self.errors) > 10:
                error_msg += f"\n... 等共 {len(self.errors)} 个错误"
            QMessageBox.warning(self, "数据警告", f"发现以下问题：\n\n{error_msg}\n\n请修正 input.txt 文件后点击「重新加载数据」按钮。")

    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("费用统计系统")
        self.setGeometry(100, 100, 1000, 700)

        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # 标题
        title_label = QLabel("📊 费用统计报表")
        if PYQT_VERSION == 5:
            title_label.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        else:
            title_label.setFont(QFont("Microsoft YaHei", 16, QFont.Weight.Bold))
        title_label.setAlignment(AlignCenter)
        layout.addWidget(title_label)

        # 统计摘要（顶部状态栏）
        self.summary_label = QLabel()
        self.summary_label.setFont(QFont("Microsoft YaHei", 10))
        self.summary_label.setAlignment(AlignCenter)
        layout.addWidget(self.summary_label)
        self.refresh_summary_label()  # 初始化摘要标签

        # 选项卡部件
        tab_widget = QTabWidget()
        layout.addWidget(tab_widget)

        # Tab 1: 详细表格
        table_tab = QWidget()
        table_layout = QVBoxLayout(table_tab)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        table_layout.addWidget(self.table)

        tab_widget.addTab(table_tab, "详细表格")

        # Tab 2: 统计摘要
        summary_tab = QWidget()
        summary_layout = QVBoxLayout(summary_tab)

        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_text.setFont(QFont("Consolas", 10))
        summary_layout.addWidget(self.summary_text)

        tab_widget.addTab(summary_tab, "统计摘要")

        # Tab 3: 原始数据
        raw_tab = QWidget()
        raw_layout = QVBoxLayout(raw_tab)

        self.raw_text = QTextEdit()
        self.raw_text.setReadOnly(True)
        self.raw_text.setFont(QFont("Consolas", 9))
        raw_layout.addWidget(self.raw_text)

        tab_widget.addTab(raw_tab, "原始数据")

        # Tab 4: 图表可视化
        chart_tab = QWidget()
        chart_layout = QVBoxLayout(chart_tab)

        # 创建 matplotlib 图表
        if MATPLOTLIB_AVAILABLE:
            self.figure = Figure(figsize=(8, 5))
            self.canvas = FigureCanvas(self.figure)
            chart_layout.addWidget(self.canvas)
            # 仅当有统计信息时才刷新图表
            if self.stats is not None and self.stats.all_dates:
                self.refresh_chart()
            else:
                # 显示空图表提示
                self.figure.clear()
                ax = self.figure.add_subplot(111)
                ax.text(0.5, 0.5, "暂无数据\n请选择费用文件",
                        horizontalalignment='center', verticalalignment='center',
                        transform=ax.transAxes, fontsize=14)
                ax.set_xticks([])
                ax.set_yticks([])
                self.canvas.draw()
        else:
            chart_label = QLabel("图表功能需要 matplotlib 库。\n请运行：pip install matplotlib")
            chart_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chart_layout.addWidget(chart_label)

        tab_widget.addTab(chart_tab, "图表可视化")

        # ========== 文件选择区域 ==========
        file_layout = QHBoxLayout()

        # 费用文件选择
        self.select_input_btn = QPushButton("📂 选择费用文件")
        self.select_input_btn.clicked.connect(self.select_input_file)
        file_layout.addWidget(self.select_input_btn)

        self.input_path_label = QLabel(os.path.basename(self.input_path) if self.input_path else "未选择")
        self.input_path_label.setStyleSheet("color: #666; font-style: italic;")
        file_layout.addWidget(self.input_path_label)

        file_layout.addSpacing(20)

        # 编号名字文件选择
        self.select_mapping_btn = QPushButton("📂 选择编号名字文件")
        self.select_mapping_btn.clicked.connect(self.select_mapping_file)
        file_layout.addWidget(self.select_mapping_btn)

        self.mapping_path_label = QLabel(os.path.basename(self.mapping_path) if self.mapping_path else "未选择")
        self.mapping_path_label.setStyleSheet("color: #666; font-style: italic;")
        file_layout.addWidget(self.mapping_path_label)

        layout.addLayout(file_layout)

        # 按钮区域
        button_layout = QHBoxLayout()

        self.reload_btn = QPushButton("🔄 重新加载数据")
        self.reload_btn.clicked.connect(self.reload_data)
        button_layout.addWidget(self.reload_btn)

        self.export_btn = QPushButton("📄 导出报告")
        self.export_btn.clicked.connect(self.export_report)
        button_layout.addWidget(self.export_btn)

        self.export_excel_btn = QPushButton("📊 导出 Excel")
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.export_excel_btn.setEnabled(OPENPYXL_AVAILABLE)
        if not OPENPYXL_AVAILABLE:
            self.export_excel_btn.setToolTip("需要安装 openpyxl: pip install openpyxl")
        button_layout.addWidget(self.export_excel_btn)

        self.copy_btn = QPushButton("📋 复制表格")
        self.copy_btn.clicked.connect(self.copy_table_to_clipboard)
        button_layout.addWidget(self.copy_btn)

        button_layout.addStretch()

        # 收集需要在加载时禁用的按钮
        self.action_buttons = [
            self.select_input_btn,
            self.select_mapping_btn,
            self.reload_btn,
            self.export_btn,
            self.export_excel_btn,
            self.copy_btn
        ]

        layout.addLayout(button_layout)

        # 注意：不在这里填充内容，由 __init__ 根据数据状态统一处理


    def refresh_table(self):
        """刷新表格"""
        # 列：编号、名称、各日期费用、总计（已删除"每日小计"列）
        cols = 2 + len(self.stats.all_dates) + 1  # 编号 + 名称 + 每日 + 总计
        data_rows = len(self.stats.num_totals)
        total_rows = data_rows + 1  # 每日总计行 + 数据行

        self.table.setRowCount(total_rows)
        self.table.setColumnCount(cols)

        # 表头
        headers = ["编号", "名称"]
        for date in self.stats.all_dates:
            headers.append(date)
        headers.append("总计")  # 只保留总计列
        self.table.setHorizontalHeaderLabels(headers)

        # ===== 第0行：每日总计行（放在最前面） =====
        label_item = QTableWidgetItem("每日总计")
        label_item.setFont(QFont("Microsoft YaHei", 9, QFont.Weight.Bold if PYQT_VERSION == 6 else QFont.Bold))
        self.table.setItem(0, 0, label_item)
        self.table.setSpan(0, 0, 1, 2)  # 合并前两列（编号+名称）

        # 填充每日总计数据（各日期列）
        for col_idx, date in enumerate(self.stats.all_dates):
            daily_total = self.stats.daily_totals[date]
            item = QTableWidgetItem(f"{daily_total:.1f}")
            item.setTextAlignment(AlignCenter)
            item.setFont(QFont("Microsoft YaHei", 9, QFont.Weight.Bold if PYQT_VERSION == 6 else QFont.Bold))
            self.table.setItem(0, 2 + col_idx, item)

        # 总计列（总总计）
        grand_total_item = QTableWidgetItem(f"{self.stats.grand_total:.1f}")
        grand_total_item.setTextAlignment(AlignCenter)
        grand_total_item.setFont(QFont("Microsoft YaHei", 9, QFont.Weight.Bold if PYQT_VERSION == 6 else QFont.Bold))
        self.table.setItem(0, 2 + len(self.stats.all_dates), grand_total_item)

        # ===== 第1行开始：各编号数据 =====
        for row_idx, num in enumerate(sorted(self.stats.num_totals.keys()), start=1):
            name = self.num_to_name.get(num, f"编号{num}")
            daily_costs = [self.stats.daily_data[date].get(num, 0.0) for date in self.stats.all_dates]
            total = self.stats.num_totals[num]

            # 编号
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(num)))
            # 名称
            self.table.setItem(row_idx, 1, QTableWidgetItem(name))

            # 每日费用
            for col_idx, cost in enumerate(daily_costs):
                item = QTableWidgetItem(f"{cost:.1f}")
                item.setTextAlignment(AlignCenter)
                self.table.setItem(row_idx, 2 + col_idx, item)

            # 总计
            item = QTableWidgetItem(f"{total:.1f}")
            item.setTextAlignment(AlignCenter)
            self.table.setItem(row_idx, 2 + len(self.stats.all_dates), item)

        # 调整列宽
        if PYQT_VERSION == 5:
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        else:
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        # 额外：确保日期列有足够宽度
        for col in range(2, 2 + len(self.stats.all_dates)):
            if PYQT_VERSION == 5:
                self.table.setColumnWidth(col, 80)  # 日期列固定宽度
            else:
                self.table.setColumnWidth(col, 80)
        self.table.horizontalHeader().setStretchLastSection(True)

    def refresh_summary(self):
        """刷新统计摘要"""
        report = StatisticsEngine.generate_report(
            self.records, self.stats, self.num_to_name
        )
        self.summary_text.setPlainText(report)

    def refresh_raw_data(self):
        """刷新原始数据显示"""
        lines = []
        lines.append("=" * 60)
        lines.append("原始数据（input.txt）")
        lines.append("=" * 60)
        for record in self.records:
            items_str = " ".join([f"[{num},{cost}]" for num, cost in record.items])
            lines.append(f"{record.date_str}: {items_str}")
        self.raw_text.setPlainText("\n".join(lines))

    def refresh_summary_label(self):
        """刷新顶部统计摘要标签"""
        if self.stats and self.stats.all_dates:
            total = self.stats.grand_total
            days = len(self.stats.all_dates)
            nums = len(self.stats.num_totals)
            self.summary_label.setText(f"总费用：{total:.1f} 元 | 统计日期：{days} 天 | 编号数量：{nums} 个")
        else:
            if not self.input_path:
                self.summary_label.setText("未选择费用文件 | 请点击「选择费用文件」按钮")
            elif self.errors:
                self.summary_label.setText("文件格式错误 | 请重新选择文件")
            else:
                self.summary_label.setText("暂无数据 | 请选择费用文件")

    def show_no_data_state(self):
        """显示无数据状态（当数据解析失败或未选择文件时调用）"""
        # 表格：显示空表格或提示
        self.table.setRowCount(1)
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["状态"])

        # 根据状态显示不同提示
        if not self.input_path:
            status_text = "请选择费用文件\n点击上方「选择费用文件」按钮"
        elif self.errors:
            status_text = "文件格式有误\n请重新选择正确的文件"
        else:
            status_text = "暂无有效数据\n请检查文件格式"

        item = QTableWidgetItem(status_text)
        item.setTextAlignment(AlignCenter)
        self.table.setItem(0, 0, item)

        # 摘要：显示状态信息
        summary_lines = []
        summary_lines.append("=" * 60)
        summary_lines.append("费用统计报告")
        summary_lines.append("=" * 60)
        summary_lines.append("")

        if not self.input_path:
            summary_lines.append("【等待选择文件】")
            summary_lines.append("-" * 60)
            summary_lines.append("  尚未选择费用文件。")
            summary_lines.append("  请点击上方「选择费用文件」按钮选择 input.txt 或费用数据文件。")
        elif self.errors:
            summary_lines.append("【文件错误】")
            summary_lines.append("-" * 60)
            for err in self.errors:
                summary_lines.append(f"  • {err}")
            summary_lines.append("")
            summary_lines.append("请修正文件后，重新点击「重新加载数据」或重新选择文件。")
        else:
            summary_lines.append("【无数据】")
            summary_lines.append("-" * 60)
            summary_lines.append("  文件已选择但未解析到有效数据。")
            summary_lines.append("  请检查文件格式是否符合要求。")

        summary_lines.append("=" * 60)
        self.summary_text.setPlainText("\n".join(summary_lines))

        # 原始数据：显示状态
        raw_lines = []
        raw_lines.append("=" * 60)
        raw_lines.append("原始数据状态")
        raw_lines.append("=" * 60)
        if not self.input_path:
            raw_lines.append("尚未选择费用文件")
        elif self.errors:
            raw_lines.append("文件包含以下错误：")
            for err in self.errors:
                raw_lines.append(f"  ✗ {err}")
        else:
            raw_lines.append("暂无有效数据")
        self.raw_text.setPlainText("\n".join(raw_lines))

        # 图表：显示提示
        if MATPLOTLIB_AVAILABLE and self.figure:
            self.figure.clear()
            ax = self.figure.add_subplot(111)
            if not self.input_path:
                msg = '请选择费用文件'
            elif self.errors:
                msg = '文件格式有误\n请重新选择'
            else:
                msg = '暂无数据'
            ax.text(0.5, 0.5, msg,
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
            ax.set_xticks([])
            ax.set_yticks([])
            self.canvas.draw()

    def select_input_file(self):
        """选择输入文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择输入文件", os.path.dirname(self.input_path),
            "Text Files (*.txt);;All Files (*)"
        )
        if file_path:
            self.input_path = file_path
            self.input_path_label.setText(os.path.basename(file_path))
            self.reload_data()

    def select_mapping_file(self):
        """选择编号映射文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择编号映射文件", os.path.dirname(self.mapping_path),
            "Text Files (*.txt);;All Files (*)"
        )
        if file_path:
            self.mapping_path = file_path
            self.mapping_path_label.setText(os.path.basename(file_path))
            self.reload_data()

    def refresh_chart(self):
        """刷新图表"""
        if not MATPLOTLIB_AVAILABLE:
            return

        self.figure.clear()
        ax = self.figure.add_subplot(111)

        # 配置中文字体（解决显示为框的问题）
        # 优先使用 Windows 系统常见的中文字体
        chinese_fonts = ['SimHei', 'Microsoft YaHei', 'PingFang SC', 'Hiragino Sans GB', 'Heiti TC', 'STHeiti']
        font_found = False
        for font_name in chinese_fonts:
            try:
                import matplotlib.font_manager as fm
                # 检查字体是否可用
                if any(font_name.lower() in f.name.lower() for f in fm.fontManager.ttflist):
                    plt.rcParams['font.sans-serif'] = [font_name]
                    plt.rcParams['axes.unicode_minus'] = False  # 正确显示负号
                    font_found = True
                    break
            except:
                continue

        # 备用方案：设置通用中文字体配置
        if not font_found:
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'PingFang SC', 'Hiragino Sans GB', 'Heiti TC', 'STHeiti', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False

        # 准备数据
        dates = self.stats.all_dates
        totals = [self.stats.daily_totals[date] for date in dates]

        # 转换日期格式用于显示
        date_labels = [date.replace('-', '/') for date in dates]

        # 绘制折线图
        ax.plot(date_labels, totals, marker='o', linewidth=2, markersize=6, color='#2E86AB')
        ax.fill_between(date_labels, totals, alpha=0.3, color='#2E86AB')

        # 设置标签（字体已在上面配置）
        ax.set_xlabel('日期', fontsize=12)
        ax.set_ylabel('费用（元）', fontsize=12)
        ax.set_title('每日费用趋势图', fontsize=14, fontweight='bold')

        # 旋转 x 轴标签
        ax.tick_params(axis='x', rotation=45)

        # 网格
        ax.grid(True, alpha=0.3, linestyle='--')

        # 在每个点上显示数值
        for i, (date, total) in enumerate(zip(date_labels, totals)):
            ax.annotate(f'{total:.1f}', xy=(date, total), xytext=(0, 10),
                       textcoords='offset points', ha='center', fontsize=9)

        self.figure.tight_layout()
        self.canvas.draw()

    def reload_data(self):
        """重新加载数据（使用当前选择的文件）- 异步版本"""
        # 验证输入文件是否已选择
        if not self.input_path:
            QMessageBox.warning(self, "提示", "请先选择费用文件！")
            return

        # 如果已在加载中，忽略（或可取消上次）
        if self.is_loading:
            QMessageBox.information(self, "提示", "正在加载中，请稍候...")
            return

        # 如果未选择映射文件，给出提示但仍继续（映射文件是可选的）
        if not self.mapping_path:
            QMessageBox.information(self, "提示", "未选择编号名字文件，将只显示编号数字。\n如需显示名称，请选择编号名字文件。")

        # 启动后台线程
        self.start_loading()

        # 创建并启动 worker 线程
        self.loading_thread = QThread()
        self.loading_worker = DataLoaderWorker(self.input_path, self.mapping_path)
        self.loading_worker.moveToThread(self.loading_thread)

        # 连接信号
        self.loading_thread.started.connect(self.loading_worker.run)
        self.loading_worker.finished_signal.connect(self.on_data_loaded)
        self.loading_worker.error_signal.connect(self.on_loading_error)
        # 线程结束时清理
        self.loading_worker.finished_signal.connect(self.loading_thread.quit)
        self.loading_worker.error_signal.connect(self.loading_thread.quit)
        self.loading_thread.finished.connect(self.on_loading_finished)

        # 启动
        self.loading_thread.start()

    def export_report(self):
        """导出报告到 txt 文件"""
        try:
            report = StatisticsEngine.generate_report(
                self.records, self.stats, self.num_to_name
            )

            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存报告", f"费用统计报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                "Text Files (*.txt);;All Files (*)"
            )

            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(report)
                QMessageBox.information(self, "成功", f"报告已保存到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{str(e)}")

    def export_excel(self):
        """导出表格数据到 Excel 文件"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                self, "库缺失",
                "Excel 导出需要 openpyxl 库。\n请运行：pip install openpyxl"
            )
            return

        try:
            # 创建 Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "费用统计"

            # 定义样式
            header_font = Font(name='Microsoft YaHei', bold=True, size=11)
            daily_total_font = Font(name='Microsoft YaHei', bold=True, size=10)
            data_font = Font(name='Consolas', size=10)
            total_font = Font(name='Microsoft YaHei', bold=True, size=10, color='FF0000')

            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')

            # 浅色背景填充
            header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
            daily_total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            total_fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ========== 写入表头 ==========
            headers = ["编号", "名称"] + [date for date in self.stats.all_dates] + ["总计"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = center_alignment
                cell.fill = header_fill
                cell.border = thin_border

            # ========== 写入第2行：每日总计 ==========
            # 合并前两列显示 "每日总计"
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
            label_cell = ws.cell(row=2, column=1, value="每日总计")
            label_cell.font = daily_total_font
            label_cell.alignment = left_alignment
            label_cell.fill = daily_total_fill
            label_cell.border = thin_border
            # 合并区域的边框需要单独设置
            for col in [1, 2]:
                ws.cell(row=2, column=col).border = thin_border

            # 写入每日总计数据（各日期列）
            for col_idx, date in enumerate(self.stats.all_dates, start=3):
                daily_total = self.stats.daily_totals[date]
                cell = ws.cell(row=2, column=col_idx, value=round(daily_total, 2))
                cell.font = daily_total_font
                cell.alignment = center_alignment
                cell.fill = daily_total_fill
                cell.border = thin_border
                cell.number_format = '#,##0.00'

            # 总计列（总总计）
            grand_total_cell = ws.cell(row=2, column=len(headers), value=round(self.stats.grand_total, 2))
            grand_total_cell.font = total_font
            grand_total_cell.alignment = center_alignment
            grand_total_cell.fill = total_fill
            grand_total_cell.border = thin_border
            grand_total_cell.number_format = '#,##0.00'

            # ========== 写入第3行及之后：各编号数据 ==========
            for row_idx, num in enumerate(sorted(self.stats.num_totals.keys()), start=3):
                name = self.num_to_name.get(num, f"编号{num}")
                daily_costs = [self.stats.daily_data[date].get(num, 0.0) for date in self.stats.all_dates]
                total = self.stats.num_totals[num]

                # 编号
                num_cell = ws.cell(row=row_idx, column=1, value=num)
                num_cell.font = data_font
                num_cell.alignment = center_alignment
                num_cell.border = thin_border

                # 名称
                name_cell = ws.cell(row=row_idx, column=2, value=name)
                name_cell.font = data_font
                name_cell.alignment = left_alignment
                name_cell.border = thin_border

                # 每日费用
                for col_idx, cost in enumerate(daily_costs, start=3):
                    cost_cell = ws.cell(row=row_idx, column=col_idx, value=round(cost, 2))
                    cost_cell.font = data_font
                    cost_cell.alignment = center_alignment
                    cost_cell.border = thin_border
                    cost_cell.number_format = '#,##0.00'

                # 总计
                total_cell = ws.cell(row=row_idx, column=len(headers), value=round(total, 2))
                total_cell.font = total_font
                total_cell.alignment = center_alignment
                total_cell.border = thin_border
                total_cell.number_format = '#,##0.00'

            # ========== 调整列宽 ==========
            ws.column_dimensions['A'].width = 8   # 编号
            ws.column_dimensions['B'].width = 15  # 名称
            for col_idx in range(3, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 10

            # ========== 保存文件 ==========
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存 Excel 文件",
                f"费用统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, "成功", f"Excel 文件已保存到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出 Excel 失败：\n{str(e)}")

    def copy_table_to_clipboard(self):
        """复制表格内容到剪贴板（包含表头）"""
        try:
            from PyQt5.QtWidgets import QApplication
            from PyQt6.QtWidgets import QApplication as Qt6App
            from PyQt5.QtCore import QMimeData
            from PyQt6.QtCore import QMimeData as Qt6MimeData
            from PyQt5.QtGui import QClipboard
            from PyQt6.QtGui import QClipboard as Qt6Clipboard

            # 构建 tab-separated 文本
            lines = []

            # 添加表头
            headers = ["编号", "名称"] + [date for date in self.stats.all_dates] + ["总计"]
            lines.append("\t".join(headers))

            # 添加每日总计行
            daily_total_row = ["每日总计", ""] + [f"{self.stats.daily_totals[date]:.1f}" for date in self.stats.all_dates] + [f"{self.stats.grand_total:.1f}"]
            lines.append("\t".join(daily_total_row))

            # 添加各编号数据行
            for num in sorted(self.stats.num_totals.keys()):
                name = self.num_to_name.get(num, f"编号{num}")
                daily_costs = [f"{self.stats.daily_data[date].get(num, 0.0):.1f}" for date in self.stats.all_dates]
                total = f"{self.stats.num_totals[num]:.1f}"
                row = [str(num), name] + daily_costs + [total]
                lines.append("\t".join(row))

            # 复制到剪贴板
            clipboard_text = "\n".join(lines)

            # 获取剪贴板对象（兼容 PyQt5/6）
            if PYQT_VERSION == 5:
                clipboard = QApplication.clipboard()
                mime_data = QMimeData()
            else:
                clipboard = Qt6App.clipboard()
                mime_data = Qt6MimeData()

            mime_data.setText(clipboard_text)
            clipboard.setMimeData(mime_data)

            QMessageBox.information(self, "成功", "表格已复制到剪贴板！\n可以直接粘贴到 Excel 中。")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"复制失败：\n{str(e)}")

    def start_loading(self):
        """开始加载状态"""
        self.is_loading = True
        # 禁用所有操作按钮
        for btn in self.action_buttons:
            btn.setEnabled(False)
        # 更新顶部标签显示加载中
        self.summary_label.setText("正在加载数据，请稍候...")
        # 显示加载占位界面
        self.show_loading_state()

    def show_loading_state(self):
        """显示加载中的占位界面"""
        self.table.setRowCount(1)
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["状态"])
        item = QTableWidgetItem("正在加载数据，请稍候...")
        item.setTextAlignment(AlignCenter)
        self.table.setItem(0, 0, item)
        summary_lines = ["=" * 60, "费用统计报告", "=" * 60, "", "正在加载数据，请稍候...", "=" * 60]
        self.summary_text.setPlainText("\n".join(summary_lines))
        self.raw_text.setPlainText("正在加载数据...")
        if MATPLOTLIB_AVAILABLE and self.figure:
            self.figure.clear()
            ax = self.figure.add_subplot(111)
            ax.text(0.5, 0.5, "正在加载数据...",
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
            ax.set_xticks([])
            ax.set_yticks([])
            self.canvas.draw()

    def on_data_loaded(self, records, stats, num_to_name, errors, input_path, mapping_path):
        """后台加载完成后的UI更新"""
        self.records = records
        self.stats = stats
        self.num_to_name = num_to_name
        self.errors = errors
        self.input_path = input_path
        self.mapping_path = mapping_path
        self.is_loading = False

        if errors:
            error_msg = "\n".join([f"• {err}" for err in errors[:10]])
            if len(errors) > 10:
                error_msg += f"\n... 等共 {len(errors)} 个错误"
            QMessageBox.warning(self, "数据警告", f"发现以下格式问题：\n\n{error_msg}\n\n请修正文件后重新选择。")
            self.records = []
            self.stats = Statistics(daily_data={}, num_totals={}, daily_totals={}, grand_total=0.0, all_dates=[])
            self.num_to_name = {}
            self.show_no_data_state()
            self.refresh_summary_label()
            return

        if not records:
            QMessageBox.warning(self, "警告", "未解析到任何有效数据！\n请检查费用文件格式是否正确。")
            self.records = []
            self.stats = Statistics(daily_data={}, num_totals={}, daily_totals={}, grand_total=0.0, all_dates=[])
            self.num_to_name = {}
            self.errors = []
            self.show_no_data_state()
            self.refresh_summary_label()
            return

        self.errors = []
        self.refresh_table()
        self.refresh_summary()
        self.refresh_raw_data()
        if MATPLOTLIB_AVAILABLE:
            self.refresh_chart()
        self.refresh_summary_label()
        QMessageBox.information(self, "成功", f"数据已重新加载！\n共 {len(records)} 天，{len(stats.num_totals)} 个编号")

    def on_loading_error(self, error_msg):
        """加载过程中出现异常"""
        self.is_loading = False
        QMessageBox.critical(self, "错误", f"加载数据时发生异常：\n{error_msg}")
        self.records = []
        self.stats = Statistics(daily_data={}, num_totals={}, daily_totals={}, grand_total=0.0, all_dates=[])
        self.num_to_name = {}
        self.errors = [f"系统错误: {error_msg}"]
        self.show_no_data_state()
        self.refresh_summary_label()

    def on_loading_finished(self):
        """线程结束后的清理"""
        self.is_loading = False
        # 重新启用所有按钮
        for btn in self.action_buttons:
            btn.setEnabled(True)
        self.loading_thread = None
        self.loading_worker = None



# ==================== 主程序入口 ====================

def main():
    """主函数"""
    # 初始状态：不自动加载文件，等待用户选择
    records = []
    num_to_name = {}
    stats = Statistics(
        daily_data={}, num_totals={}, daily_totals={}, grand_total=0.0, all_dates=[]
    )
    errors = []  # 初始无错误，因为没有尝试加载

    # 启动 GUI（传递空路径，表示尚未选择文件）
    app = QApplication(sys.argv)
    window = StatisticsGUI(records, stats, num_to_name, errors,
                          input_path="",  # 空路径表示未选择
                          mapping_path="")  # 空路径表示未选择
    window.show()
    if PYQT_VERSION == 5:
        sys.exit(app.exec_())
    else:
        sys.exit(app.exec())


if __name__ == "__main__":
    main()
